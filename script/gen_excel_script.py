# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @createTime: 19-1-23 上午9:35
# @author: scdev030
from openpyxl import load_workbook, Workbook
import random
import time
import uuid
import datetime


class GenExcelScript(object):
    """
    生成excel文件
    """

    def __init__(self):
        super(GenExcelScript, self).__init__()
        self.table_name = u'测试sheet1'
        self.excel_file_name = u'测试表格.xlsx'

    def run(self):
        """
        控制运行
        """
        table_head_list = self._gen_table_head_list()
        content_list_list = self._gen_content_list_list()
        self._run_gen_excel(table_head_list, content_list_list)

    def _run_gen_excel(self, table_head_list, content_list_list):
        """
        生成表格
        """
        try:
            wb = load_workbook(self.excel_file_name)
        except IOError:
            wb = Workbook()
        wb.create_sheet(self.table_name)
        ws = wb.get_sheet_by_name(self.table_name)
        ws.append(table_head_list)
        for one_list in content_list_list:
            ws.append(one_list)
        try:
            # 删掉默认生成的sheet
            del wb['Sheet']
        except KeyError:
            pass
        wb.save(self.excel_file_name)
        wb.close()

    @staticmethod
    def _gen_content_list_list():
        """
        生成写入内容,一行一个list和表头对应
        """
        content_list_list = list()
        for i in range(1, 100):
            tmp_list = list()
            tmp_list.append(int(time.time()))
            tmp_list.append(str(uuid.uuid4()))
            tmp_list.append(str(random.random()))
            tmp_list.append(datetime.datetime.now().strftime("%Y-%m-%d"))
            content_list_list.append(tmp_list)
        return content_list_list

    @staticmethod
    def _gen_table_head_list():
        """
        表头
        """
        table_head_list = [u'时间戳', u'uuid4', u'随机数', u'今日日期']
        return table_head_list


if __name__ == '__main__':
    worker = GenExcelScript()
    worker.run()
